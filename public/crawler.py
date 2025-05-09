import threading
import queue
import datetime
import pyautogui
import pyperclip
import time
import openpyxl as op
from selenium import webdriver
from selenium.webdriver.common.by import By

# 카페 주소
free_talk = 'https://cafe.naver.com/skybluezw4rh?iframe_url=/ArticleList.nhn%3Fsearch.clubid=29434212%26search.boardtype=L%26search.menuid=16%26search.marketBoardTab=D%26search.specialmenutype=%26userDisplay=50'
all_post = 'https://cafe.naver.com/ArticleList.nhn?search.clubid=29434212&search.boardtype=L&search.menuid=&search.marketBoardTab=D&search.specialmenutype=&userDisplay=50'
# 큐 생성 -> 쓰레드 사이에서 데이터 공유할때 안전하게 하려면 큐를 써야한다고 함
_data_queue = queue.Queue() 
filtered_data_queue = queue.Queue()

class DataCollector(threading.Thread):
    def __init__(self, driver, output_queue):
        super().__init__()
        self.driver = driver
        self.output_queue = output_queue

    def run(self):
        while True:
            try:
                self.driver.get(all_post)
                time.sleep(1)
                self.driver.switch_to.frame('cafe_main')

                table = self.driver.find_element(By.XPATH, '/html/body/div[1]/div/div[4]')
                tbody = table.find_element(By.TAG_NAME, 'tbody')
                rows = tbody.find_elements(By.TAG_NAME, 'tr')
                valid_rows = [row for row in rows if row.find_elements(By.CLASS_NAME, 'article')]

                data_list = []
                for index, row in enumerate(valid_rows):
                    try:
                        # heading = row.find_element(By.CLASS_NAME, 'head').text # 말머리 -> 전체글보기에는 없음
                        heading = row.find_element(By.CLASS_NAME,'board-name').text.strip()
                    except:
                        # heading = '말머리 없음'
                        heading = '게시판 확인 불가'
                    title = row.find_element(By.CLASS_NAME, 'article').text
                    if heading in title:
                        title = title.replace(heading, '').strip()
                    try:
                        comment = row.find_element(By.CLASS_NAME,'cmt').text.replace('\n','')
                    except:
                        comment = '댓글 없음'
                    name = row.find_element(By.CLASS_NAME, 'td_name').text
                    date = row.find_element(By.CLASS_NAME, 'td_date').text
                    view = row.find_element(By.CLASS_NAME, 'td_view').text
                    try:
                        likes = row.find_element(By.CLASS_NAME, 'td_likes').text
                    except:
                        likes = '알 수 없음' # 좋아요는 전체글보기에는 없음
                    row_data = [index + 1, heading, title,comment, name, date, view, likes]
                    
                    data_list.append(row_data)
                self.output_queue.put(data_list)
                print('데이터 큐에 저장 완료')
            except Exception as e:
                print('수집 오류:', e)
            time.sleep(10)  # 20초 대기
            self.driver.refresh()  # 페이지 새로고침
            time.sleep(3)
            print('페이지 새로고침 완료')

class DataFilter(threading.Thread):
    
    def __init__(self, input_queue, output_queue):
        super().__init__()
        self.input_queue = input_queue
        self.output_queue = output_queue
        self.prev_data = []

    def run(self):
        while True:
            data = self.input_queue.get()
            filtered = [row for row in data if row not in self.prev_data]
            if filtered:
                self.prev_data.extend(filtered)
                self.output_queue.put(filtered)
                print(f'필터링 완료 {len(filtered)}개 새로운 데이터')
            else:
                print('필터링 실패 or 중복 데이터 없음')
                time.sleep(1)
            # 필터링 완료 된 데이터 큐에 저장
            self.input_queue.task_done()

class DataSaveExcel(threading.Thread):
    
    def __init__(self, input_queue):
        super().__init__()
        # 맨 처음에 엑셀 파일 하나 생성
        self.input_queue = input_queue
        self.curr_date = datetime.datetime.today()
        self._create_new_workbook() # 생성날짜 기록 o

    def _create_new_workbook(self):
        self.wb = op.Workbook()
        self.create_time = datetime.datetime.today() # 생성날짜 기록
        self.save_time = self.curr_date.strftime('%Y%m%d_%H%M%S') # 저장용 파일명에 쓸 거
        self.file_path = fr'D:\Kim goeun\database\mom_cafe_{self.save_time}.xlsx'
        self.wb.save(self.file_path) # 엑셀 파일 저장
        print('엑셀 파일 생성 완료')

    def run(self):
        while True:
            #추가할 데이터 큐에서 가져옴
            data = self.input_queue.get()
            #now = 현재 시간, today = 오늘 날짜
            now = datetime.datetime.now()
            sheet_name = now.strftime('%H시')
            
            try:
                if now.date() != self.create_time.date(): # 날짜가 바뀌면 엑셀 파일 새로 생성
                    self._create_new_workbook()
                    self.wb = op.load_workbook(self.file_path)
                else:
                    self.wb = op.load_workbook(self.file_path)
            except Exception as e:
                print('엑셀 파일 로드 오류:', e)
                self._create_new_workbook()
                self.wb = op.load_workbook(self.file_path)

            if sheet_name not in self.wb.sheetnames:
                ws = self.wb.create_sheet(title=sheet_name)
                ws.append(['번호', '게시판', '글 제목', '댓글 수', '작성자', '작성일', '조회수', '좋아요'])
   
            ws = self.wb[sheet_name]

            for row in data:
                ws.append(row)
                        
            self.wb.save(self.file_path) # 엑셀 파일 저장
            self.wb.close()
            self.input_queue.task_done()
            time.sleep(1)

# 메인
def main():
    
    # 네이버 로그인
    login_page = 'https://nid.naver.com/nidlogin.login?mode=form&url=https://www.naver.com/' # 네이버 로그인 페이지
    query = ['**','**'] # id, pw
    css=['id','pw']
    
    driver = webdriver.Chrome()    
    driver.get(login_page)
    
    try:
        for i in range(len(query)):
            driver.find_element(By.ID,css[i]).click()
            time.sleep(1)
            pyperclip.copy(query[i])
            pyautogui.hotkey('ctrl','v')
            time.sleep(2)
        print('로그인 정보 입력 완료')    
        pyautogui.hotkey('Enter')
        time.sleep(1.5)
        driver.find_element(By.ID,'new.dontsave').click()
        print('로그인 성공')
        print('-' * 40)
        time.sleep(2)
    except Exception as e:
        print('로그인 오류:', e)
    
    collector = DataCollector(driver, _data_queue)
    filterer = DataFilter(_data_queue, filtered_data_queue)
    saver = DataSaveExcel(filtered_data_queue)

    collector.start()
    filterer.start()
    saver.start()

    collector.join()
    filterer.join()
    saver.join()

if __name__ == '__main__':
    main()

'''
1. 엑셀로 저장해 둔 뒤 조회수, 댓글 수 확인
2. 반응이 좋은 글 패턴 분석 가능 -> 제목에 자주 등장하는 단어들 (ex.후기, ~맘들 ~ ? 하는 질문글 등) 이런거 찾아서 마케팅 문구에 활용
3. 댓글 수 -> 이슈가 되는 소재일 가능성 높음 -> 저만 ~ 별로라고 생각되나요? , 다들 ~ 어떤거 쓰세요? 하는 글들 -> 이슈가 되는 키워드 찾기 쉬움
4. 잠입 마케팅 -> 제목으로 흥미를 끌어야 조회수가 높아지니 활용 가능성 매우 높음

5. 나중에 글 + 댓글 내용또한 수집하면 더 정확한 분석 가능 + 마케팅이라는걸 들키지 않고 더 자연스러운 홍보 가능
'''
'''

            self.wb.remove(self.wb['Sheet'])  # 기본 시트 삭제
            print('기본 시트 삭제 완료')
            
    '''